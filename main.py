#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
工会固定资产管理系统 V18.0
单文件整合版 - 财政部《工会会计制度》合规版

版本：V18.0
日期：2026-03-27
修改内容：
1. 完善导入模板：在 Excel 填写说明中完整列出所有资产类别
2. 基于V17.0升级，折旧精准计算逻辑完全保留
3. 所有功能不变，仅优化模板说明内容

历史版本：
V17.0 - 最终精准修复折旧计算
V16.0 - 最终修复折旧差额计算错误
V15.0 - 修复折旧差额计算
V14.0 - 修复语法错误
V13.0 - 修复自动编码BUG，按 GB/T 14885-2022 自动生成对应类别国标编码
V12.0 - 修复导入提示和trace警告
V11.0 - 按《工会会计制度》重构折旧计算引擎
V10.8 - 工会资产当月计提版本

编码规则：GB/T 14885-2022 国标码-YY-MM-DD-XXX
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
import hashlib
import time
import os
from datetime import datetime
from dataclasses import dataclass, asdict, field
from typing import Dict, List, Optional, Tuple
from collections import defaultdict
import json

# 尝试导入 openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


# ==================== 常量配置 ====================

VERSION = "18.0"
DB_PATH = os.path.join(os.path.dirname(__file__), 'assets.db')

# 折旧年限配置（月）- 符合工会会计制度
DEPRECIATION_YEARS_MAP = {
    '房屋及构筑物 - 钢结构': 600,
    '房屋及构筑物 - 钢筋混凝土结构': 600,
    '房屋及构筑物 - 砖混结构': 360,
    '房屋及构筑物 - 砖木结构': 360,
    '房屋及构筑物 - 简易房': 96,
    '房屋及构筑物 - 房屋附属设施': 96,
    '房屋及构筑物 - 构筑物': 96,
    '通用设备 - 计算机设备': 72,
    '通用设备 - 办公设备': 72,
    '通用设备 - 车辆': 96,
    '通用设备 - 图书档案设备': 60,
    '通用设备 - 机械设备': 120,
    '通用设备 - 电气设备': 60,
    '通用设备 - 雷达、无线电和卫星导航设备': 120,
    '通用设备 - 通信设备': 60,
    '通用设备 - 广播、电视、电影设备': 60,
    '通用设备 - 仪器仪表': 60,
    '通用设备 - 电子和通信测量设备': 60,
    '通用设备 - 计量标准器具及量具、衡器': 60,
    '专用设备 - 食品加工专用设备': 120,
    '专用设备 - 纺织设备': 120,
    '专用设备 - 缝纫、服饰、制革和毛皮加工设备': 120,
    '专用设备 - 医疗设备': 60,
    '专用设备 - 安全生产设备': 120,
    '专用设备 - 环境污染防治设备': 120,
    '专用设备 - 文艺设备': 60,
    '专用设备 - 体育设备': 60,
    '专用设备 - 娱乐设备': 60,
    '家具、用具及装具 - 家具': 180,
    '家具、用具及装具 - 用具、装具': 60,
}

CATEGORY_ALIAS = {
    '计算机设备': '通用设备 - 计算机设备',
    '办公设备': '通用设备 - 办公设备',
    '专用设备': '专用设备 - 其他',
    '房屋及构筑物': '房屋及构筑物 - 钢筋混凝土结构',
    '车辆': '通用设备 - 车辆',
    '家具': '家具、用具及装具 - 家具',
}

# GB/T 14885-2022 固定资产等资产基础分类与代码
# 格式：大类(2位) + 中类(2位) + 小类(2位) + 品目(2位)
GB_T_14885_CODES = {
    # 房屋及构筑物 (01)
    '房屋及构筑物 - 钢结构': 'A01010101',
    '房屋及构筑物 - 钢筋混凝土结构': 'A01010102',
    '房屋及构筑物 - 砖混结构': 'A01010103',
    '房屋及构筑物 - 砖木结构': 'A01010104',
    '房屋及构筑物 - 简易房': 'A01010199',
    '房屋及构筑物 - 房屋附属设施': 'A01010201',
    '房屋及构筑物 - 构筑物': 'A01010301',
    # 通用设备 (02)
    '通用设备 - 计算机设备': 'A02010101',
    '通用设备 - 办公设备': 'A02010201',
    '通用设备 - 车辆': 'A02010301',
    '通用设备 - 图书档案设备': 'A02010401',
    '通用设备 - 机械设备': 'A02010501',
    '通用设备 - 电气设备': 'A02010601',
    '通用设备 - 雷达、无线电和卫星导航设备': 'A02010701',
    '通用设备 - 通信设备': 'A02010801',
    '通用设备 - 广播、电视、电影设备': 'A02010901',
    '通用设备 - 仪器仪表': 'A02011001',
    '通用设备 - 电子和通信测量设备': 'A02011101',
    '通用设备 - 计量标准器具及量具、衡器': 'A02011201',
    # 专用设备 (03)
    '专用设备 - 食品加工专用设备': 'A03010101',
    '专用设备 - 纺织设备': 'A03010201',
    '专用设备 - 缝纫、服饰、制革和毛皮加工设备': 'A03010301',
    '专用设备 - 医疗设备': 'A03010401',
    '专用设备 - 安全生产设备': 'A03010501',
    '专用设备 - 环境污染防治设备': 'A03010601',
    '专用设备 - 文艺设备': 'A03010701',
    '专用设备 - 体育设备': 'A03010801',
    '专用设备 - 娱乐设备': 'A03010901',
    # 家具、用具及装具 (06)
    '家具、用具及装具 - 家具': 'A06010101',
    '家具、用具及装具 - 用具、装具': 'A06010201',
}

VALID_STATUS = ['在用', '闲置', '报废', '维修中', '待报废', '待维修']
USER_CREDENTIALS = {
    'pm': 'pm123',
    'dev': 'dev123',
    'qa': 'qa123',
    'admin': 'admin123'
}


# ==================== 工具函数 ====================

def get_depreciation_months(category: str) -> int:
    """获取折旧年限（月）"""
    if category in DEPRECIATION_YEARS_MAP:
        return DEPRECIATION_YEARS_MAP[category]
    if category in CATEGORY_ALIAS:
        mapped = CATEGORY_ALIAS[category]
        if mapped in DEPRECIATION_YEARS_MAP:
            return DEPRECIATION_YEARS_MAP[mapped]
    return 60


def hash_password(password: str) -> str:
    """SHA-256 密码哈希"""
    return hashlib.sha256(password.encode()).hexdigest()


def generate_asset_code(purchase_date: str, category: str, existing_codes: List[str] = None) -> Tuple[str, Optional[str]]:
    """
    生成资产编码
    格式：国标分类码-YY-MM-DD-XXX
    依据：GB/T 14885-2022《固定资产等资产基础分类与代码》
    """
    warning = None
    
    try:
        dt = datetime.strptime(purchase_date, "%Y-%m-%d")
    except ValueError:
        try:
            date_part = purchase_date.split()[0]
            dt = datetime.strptime(date_part, "%Y-%m-%d")
            warning = f"日期包含时间部分，已自动提取：{date_part}"
        except (ValueError, IndexError):
            try:
                for fmt in ["%Y/%m/%d", "%Y.%m.%d"]:
                    try:
                        dt = datetime.strptime(purchase_date, fmt)
                        warning = f"日期格式已转换：{purchase_date}"
                        break
                    except ValueError:
                        continue
                else:
                    warning = f"日期格式错误：'{purchase_date}'，已使用当前日期"
                    dt = datetime.now()
            except:
                warning = f"日期解析失败，已使用当前日期"
                dt = datetime.now()
    
    yy = dt.strftime("%y")
    mm = dt.strftime("%m")
    dd = dt.strftime("%d")
    
    # 根据资产类别获取国标编码
    gb_code = GB_T_14885_CODES.get(category, 'A02010101')  # 默认使用通用设备-计算机设备编码
    
    prefix = f"{gb_code}-{yy}-{mm}-{dd}-"
    
    max_seq = 0
    if existing_codes:
        for code in existing_codes:
            if code.startswith(prefix):
                try:
                    seq = int(code.split('-')[-1])
                    max_seq = max(max_seq, seq)
                except ValueError:
                    pass
    
    new_seq = max_seq + 1
    return f"{prefix}{new_seq:03d}", warning


# ==================== 数据类 ====================

@dataclass
class FixedAsset:
    """固定资产数据类"""
    asset_id: str
    name: str
    category: str
    original_value: float
    purchase_date: str
    department: str = ""
    user: str = ""
    location: str = ""
    status: str = "在用"
    remark: str = ""
    
    @property
    def depreciation_months(self) -> int:
        return get_depreciation_months(self.category)
    
    @property
    def depreciation_years(self) -> float:
        return self.depreciation_months / 12.0
    
    def to_dict(self) -> dict:
        return asdict(self)


@dataclass
class DepreciationResult:
    """
    折旧计算结果 V11.0 - 符合工会会计制度
    - 当月购入当月计提
    - 折旧差额当月一次性处理
    - 不留尾差、不递延、不分摊
    """
    asset_id: str
    asset_name: str
    category: str
    department: str
    user: str
    location: str
    status: str
    original_value: float
    monthly_depreciation: float  # 标准月折旧额
    accumulated_depreciation: float  # 累计折旧
    net_value: float  # 净值
    depreciation_months: int  # 总折旧月数
    depreciation_years: float
    depreciable_months: int  # 已计提月数
    depreciation_status: str  # 计提状态
    purchase_date: str
    current_month_depreciation: float = 0.00  # 本月计提折旧额（含差额）
    depreciation_diff: float = 0.00  # 折旧差额（当月一次性处理）
    audit_log: str = ""  # 审计日志


# ==================== 数据库管理器 ====================

class DatabaseManager:
    """数据库管理器 - V11.0 符合开发说明书 3.1 节"""
    
    def __init__(self, db_path: str = None):
        self.db_path = db_path or DB_PATH
        self.init_database()
    
    def init_database(self):
        """初始化数据库 - 符合开发说明书 3.1 节"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # 资产主表 (assets) - 符合开发说明书 3.1 节
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS assets (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                asset_id TEXT UNIQUE NOT NULL,
                name TEXT NOT NULL,
                category TEXT NOT NULL,
                original_value REAL NOT NULL,
                purchase_date TEXT NOT NULL,
                department TEXT DEFAULT '',
                user TEXT DEFAULT '',
                location TEXT DEFAULT '',
                status TEXT DEFAULT '在用',
                remark TEXT DEFAULT '',
                created_time TEXT,
                updated_time TEXT
            )
        ''')
        
        # 用户表 (users) - V6.0 引入
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                role TEXT DEFAULT 'user',
                created_time TEXT
            )
        ''')
        
        # 审计日志表 (audit_logs) - V11.0 引入，符合开发说明书 3.1 节
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS audit_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                operation_type TEXT NOT NULL,
                asset_id TEXT NOT NULL,
                user_id TEXT,
                old_values TEXT,
                new_values TEXT,
                operation_time TEXT
            )
        ''')
        
        # 初始化默认用户
        for username, password in USER_CREDENTIALS.items():
            try:
                cursor.execute(
                    'INSERT INTO users (username, password_hash, role, created_time) VALUES (?, ?, ?, ?)',
                    (username, hash_password(password), 'admin', datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                )
            except sqlite3.IntegrityError:
                pass
        
        conn.commit()
        conn.close()
    
    def get_connection(self) -> sqlite3.Connection:
        return sqlite3.connect(self.db_path)
    
    def get_all_assets(self) -> List[FixedAsset]:
        conn = self.get_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM assets ORDER BY asset_id')
        rows = cursor.fetchall()
        conn.close()
        
        assets = []
        for row in rows:
            assets.append(FixedAsset(
                asset_id=row[1], name=row[2], category=row[3], original_value=row[4],
                purchase_date=row[5], department=row[6], user=row[7], location=row[8],
                status=row[9], remark=row[10]
            ))
        return assets
    
    def get_asset_codes(self) -> List[str]:
        conn = self.get_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT asset_id FROM assets')
        codes = [row[0] for row in cursor.fetchall()]
        conn.close()
        return codes
    
    def add_asset(self, asset: FixedAsset) -> bool:
        conn = self.get_connection()
        cursor = conn.cursor()
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            cursor.execute('''
                INSERT INTO assets (asset_id, name, category, original_value, purchase_date,
                                   department, user, location, status, remark, created_time, updated_time)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (asset.asset_id, asset.name, asset.category, asset.original_value,
                  asset.purchase_date, asset.department, asset.user, asset.location,
                  asset.status, asset.remark, now, now))
            
            # V11.0：记录审计日志
            self.log_audit('CREATE', asset.asset_id, 'admin', None, asset.to_dict())
            
            conn.commit()
            return True
        except sqlite3.IntegrityError:
            return False
        finally:
            conn.close()
    
    def update_asset(self, asset: FixedAsset) -> bool:
        conn = self.get_connection()
        cursor = conn.cursor()
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            # 获取旧数据
            old_asset = self.get_asset_by_id(asset.asset_id)
            old_values = old_asset.to_dict() if old_asset else None
            
            cursor.execute('''
                UPDATE assets SET name=?, category=?, original_value=?, purchase_date=?,
                                  department=?, user=?, location=?, status=?, remark=?, updated_time=?
                WHERE asset_id=?
            ''', (asset.name, asset.category, asset.original_value, asset.purchase_date,
                  asset.department, asset.user, asset.location, asset.status, asset.remark, now,
                  asset.asset_id))
            
            # V11.0：记录审计日志
            if cursor.rowcount > 0:
                self.log_audit('UPDATE', asset.asset_id, 'admin', old_values, asset.to_dict())
                return True
            return False
        finally:
            conn.close()
    
    def delete_asset(self, asset_id: str) -> bool:
        conn = self.get_connection()
        cursor = conn.cursor()
        try:
            # V11.0：记录审计日志
            self.log_audit('DELETE', asset_id, 'admin', None, None)
            
            cursor.execute('DELETE FROM assets WHERE asset_id = ?', (asset_id,))
            conn.commit()
            return cursor.rowcount > 0
        finally:
            conn.close()
    
    def delete_assets_batch(self, asset_ids: List[str]) -> int:
        conn = self.get_connection()
        cursor = conn.cursor()
        try:
            # V11.0：批量记录审计日志
            for asset_id in asset_ids:
                self.log_audit('DELETE', asset_id, 'admin', None, None)
            
            placeholders = ','.join('?' * len(asset_ids))
            cursor.execute(f'DELETE FROM assets WHERE asset_id IN ({placeholders})', asset_ids)
            conn.commit()
            return cursor.rowcount
        finally:
            conn.close()
    
    def get_asset_by_id(self, asset_id: str) -> Optional[FixedAsset]:
        conn = self.get_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM assets WHERE asset_id = ?', (asset_id,))
        row = cursor.fetchone()
        conn.close()
        
        if row:
            return FixedAsset(
                asset_id=row[1], name=row[2], category=row[3], original_value=row[4],
                purchase_date=row[5], department=row[6], user=row[7], location=row[8],
                status=row[9], remark=row[10]
            )
        return None
    
    def log_audit(self, operation_type: str, asset_id: str, user_id: str, 
                  old_values: dict = None, new_values: dict = None):
        """V11.0：记录审计日志 - 符合开发说明书 3.1 节"""
        conn = self.get_connection()
        cursor = conn.cursor()
        try:
            cursor.execute('''
                INSERT INTO audit_logs (operation_type, asset_id, user_id, old_values, new_values, operation_time)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (operation_type, asset_id, user_id, 
                  json.dumps(old_values) if old_values else None,
                  json.dumps(new_values) if new_values else None,
                  datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
            conn.commit()
        finally:
            conn.close()
    
    def log_depreciation(self, result: DepreciationResult, calc_date: str):
        """V11.0：记录折旧审计日志 - 符合开发说明书 4.1 节"""
        conn = self.get_connection()
        cursor = conn.cursor()
        try:
            # 使用 audit_logs 表记录折旧操作
            dep_data = {
                'calc_date': calc_date,
                'original_value': result.original_value,
                'monthly_depreciation': result.monthly_depreciation,
                'accumulated_depreciation': result.accumulated_depreciation,
                'net_value': result.net_value,
                'current_month_depreciation': result.current_month_depreciation,
                'depreciation_diff': result.depreciation_diff,
                'depreciable_months': result.depreciable_months,
                'audit_log': result.audit_log
            }
            
            cursor.execute('''
                INSERT INTO audit_logs (operation_type, asset_id, user_id, old_values, new_values, operation_time)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', ('DEPRECIATE', result.asset_id, 'system', None, json.dumps(dep_data),
                  datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
            conn.commit()
        finally:
            conn.close()
    
    def get_audit_logs(self, asset_id: str = None, limit: int = 100) -> List[dict]:
        """V11.0：查询审计日志"""
        conn = self.get_connection()
        cursor = conn.cursor()
        try:
            if asset_id:
                cursor.execute('''
                    SELECT * FROM audit_logs WHERE asset_id = ? ORDER BY operation_time DESC LIMIT ?
                ''', (asset_id, limit))
            else:
                cursor.execute('''
                    SELECT * FROM audit_logs ORDER BY operation_time DESC LIMIT ?
                ''', (limit,))
            
            rows = cursor.fetchall()
            logs = []
            for row in rows:
                logs.append({
                    'id': row[0],
                    'operation_type': row[1],
                    'asset_id': row[2],
                    'user_id': row[3],
                    'old_values': json.loads(row[4]) if row[4] else None,
                    'new_values': json.loads(row[5]) if row[5] else None,
                    'operation_time': row[6]
                })
            return logs
        finally:
            conn.close()


# ==================== 折旧计算器 V11.0 ====================

class DepreciationCalculator:
    """
    折旧计算器 V11.0 - 符合财政部《工会会计制度》（财会〔2021〕7 号）
    
    核心规则：
    1. 折旧方法：年限平均法
    2. 净残值：0
    3. 当月增加：当月计提折旧
    4. 当月减少：当月不再计提折旧
    5. 折旧差额：变动当月一次性计入，不递延、不分摊、不遗留
    """
    
    @staticmethod
    def calculate(asset: FixedAsset, current_date: str = None, for_monthly_report: bool = False) -> DepreciationResult:
        """
        计算资产折旧（V11.0 工会会计制度合规版）- 符合开发说明书 4.1 节
        
        参数：
        - asset: 资产对象
        - current_date: 查询日期（格式：YYYY-MM-DD）
        - for_monthly_report: 是否用于月度报表（剔除不计提资产）
        
        返回：
        - DepreciationResult: 折旧计算结果
        """
        if current_date is None:
            current_date = datetime.now().strftime("%Y-%m-%d")
        
        audit_logs = []
        audit_logs.append(f"开始计算：{asset.asset_id} - {asset.name}")
        audit_logs.append(f"购置日期：{asset.purchase_date}")
        audit_logs.append(f"查询日期：{current_date}")
        
        purchase = datetime.strptime(asset.purchase_date, "%Y-%m-%d")
        current = datetime.strptime(current_date, "%Y-%m-%d")
        
        # V17.0 核心：标准月折旧额 = 原值 ÷ 折旧年限 ÷ 12 = 原值 ÷ 折旧月数
        # 必须先四舍五入，再用四舍五入后的值计算差额
        monthly_depreciation_raw = asset.original_value / asset.depreciation_months
        monthly_depreciation = round(monthly_depreciation_raw, 2)  # 四舍五入保留2位
        
        # V17.0 核心：计算折旧差额（确保总折旧 = 原值，无尾差）
        # 使用四舍五入后的月折旧计算总折旧
        # 示例：5000/72=69.444...→69.44, 69.44×72=4999.68, 差额=5000-4999.68=0.32
        total_depreciation_should = round(monthly_depreciation * asset.depreciation_months, 2)
        depreciation_diff = round(asset.original_value - total_depreciation_should, 2)
        
        audit_logs.append(f"原值：{asset.original_value:.2f} 元")
        audit_logs.append(f"折旧月数：{asset.depreciation_months} 个月")
        audit_logs.append(f"月折旧额（原始）：{monthly_depreciation_raw:.4f} 元")
        audit_logs.append(f"月折旧额（四舍五入）：{monthly_depreciation:.2f} 元")
        audit_logs.append(f"理论总折旧：{total_depreciation_should:.2f} 元")
        audit_logs.append(f"折旧差额：{depreciation_diff:.2f} 元（首月一次性处理）")
        
        # V11.0 核心修复：按"年月"判断，忽略具体日期
        # 工会会计制度：当月购入，当月计提
        purchase_year_month = (purchase.year, purchase.month)
        current_year_month = (current.year, current.month)
        
        # 计算月份差
        months_diff = (current.year - purchase.year) * 12 + (current.month - purchase.month)
        
        audit_logs.append(f"购入月份：{purchase_year_month[0]}-{purchase_year_month[1]:02d}")
        audit_logs.append(f"当前月份：{current_year_month[0]}-{current_year_month[1]:02d}")
        audit_logs.append(f"月份差：{months_diff} 个月")
        
        # V11.0 核心：工会资产当月购入当月计提
        # 根据工会会计制度，所有固定资产都是当月购入当月计提
        if purchase_year_month > current_year_month:
            # 购入月份晚于当前月份，还未到计提时间
            depreciable_months = 0
            audit_logs.append("状态：购入月份晚于当前月份，未开始计提")
        else:
            # 从购入月份到当前月份（包含购入当月）
            depreciable_months = months_diff + 1
            audit_logs.append(f"状态：当月购入当月计提，已计提月数={depreciable_months}")
        
        # 限制最大计提月数
        depreciable_months = min(depreciable_months, asset.depreciation_months)
        if depreciable_months >= asset.depreciation_months:
            audit_logs.append("状态：已提足折旧")
        
        # V11.0 新增：判断是否应计提
        should_depreciate = True
        exclude_reason = ""
        
        # 已提足折旧
        if depreciable_months >= asset.depreciation_months:
            should_depreciate = False
            exclude_reason = "已提足折旧"
        
        # 待报废、待维修状态
        if asset.status in ['待报废', '待维修']:
            should_depreciate = False
            exclude_reason = f"资产状态：{asset.status}"
        
        # 月度报表时剔除不计提资产
        if for_monthly_report and not should_depreciate:
            audit_logs.append(f"月度报表剔除：{exclude_reason}")
        
        # V11.0 核心：计算累计折旧和净值
        if depreciable_months >= asset.depreciation_months:
            # 已提足折旧
            accumulated_depreciation = round(asset.original_value, 2)
            net_value = 0.00
            audit_logs.append(f"累计折旧：{accumulated_depreciation:.2f} 元（已提足）")
            audit_logs.append(f"净值：{net_value:.2f} 元")
        else:
            # 正常计提中
            accumulated_depreciation = round(monthly_depreciation * depreciable_months, 2)
            net_value = round(asset.original_value - accumulated_depreciation, 2)
            audit_logs.append(f"累计折旧：{accumulated_depreciation:.2f} 元")
            audit_logs.append(f"净值：{net_value:.2f} 元")
        
        # V15.0 核心：本月折旧额计算
        # 本月折旧额 = 当月应计提金额（含差额）
        if not should_depreciate:
            current_month_depreciation = 0.00
            current_diff = 0.00
            audit_logs.append("本月不计提折旧")
        elif depreciable_months == 0:
            current_month_depreciation = 0.00
            current_diff = 0.00
            audit_logs.append("本月不计提折旧（未到期）")
        elif depreciable_months == 1:
            # V15.0 核心：购入首月，本月折旧额 = 正常月折旧 + 折旧差额
            current_month_depreciation = round(monthly_depreciation + depreciation_diff, 2)
            current_diff = depreciation_diff
            audit_logs.append(f"首月折旧额：{current_month_depreciation:.2f} 元（含差额 {depreciation_diff:.2f} 元）")
        else:
            # 后续月份：只计提正常月折旧
            current_month_depreciation = monthly_depreciation
            current_diff = 0.00
            audit_logs.append(f"本月折旧额：{current_month_depreciation:.2f} 元")
        
        # 判断状态
        if depreciable_months >= asset.depreciation_months:
            dep_status = "已提足"
        elif depreciable_months == 0:
            dep_status = "未开始"
        else:
            dep_status = "正常计提"
        
        audit_logs.append(f"计提状态：{dep_status}")
        audit_logs.append("计算完成")
        
        return DepreciationResult(
            asset_id=asset.asset_id,
            asset_name=asset.name,
            category=asset.category,
            department=asset.department,
            user=asset.user,
            location=asset.location,
            status=asset.status,
            original_value=asset.original_value,
            monthly_depreciation=monthly_depreciation,
            accumulated_depreciation=accumulated_depreciation,
            net_value=net_value,
            depreciation_months=asset.depreciation_months,
            depreciation_years=asset.depreciation_years,
            depreciable_months=depreciable_months,
            depreciation_status=dep_status,
            purchase_date=asset.purchase_date,
            current_month_depreciation=current_month_depreciation,
            depreciation_diff=current_diff,
            audit_log="\n".join(audit_logs)
        )
    
    @staticmethod
    def calculate_batch(assets: List[FixedAsset], current_date: str = None, 
                       for_monthly_report: bool = False) -> List[DepreciationResult]:
        """批量计算折旧 - 符合开发说明书 4.1 节"""
        results = []
        for asset in assets:
            result = DepreciationCalculator.calculate(asset, current_date, for_monthly_report)
            results.append(result)
        return results
    
    @staticmethod
    def generate_accounting_entry(results: List[DepreciationResult]) -> str:
        """
        生成会计分录（凭证）- 符合开发说明书 4.1 节
        借：资产基金——固定资产
        贷：累计折旧
        """
        total = sum(r.current_month_depreciation for r in results if r.current_month_depreciation > 0)
        
        if total == 0:
            return "无需计提折旧"
        
        entry = f"""会计分录
═══════════════════════════════════════
借：资产基金——固定资产    ¥{total:,.2f}
贷：累计折旧              ¥{total:,.2f}
═══════════════════════════════════════
合计：¥{total:,.2f}
计提资产数量：{len([r for r in results if r.current_month_depreciation > 0])} 项
"""
        return entry


# ==================== GUI 对话框 ====================

class LoginDialog(tk.Toplevel):
    """登录对话框 - 符合开发说明书 5.1 节"""
    
    def __init__(self, parent):
        super().__init__(parent)
        self.title("用户登录")
        self.geometry("350x200")
        self.result = None
        self.transient(parent)
        self.grab_set()
        
        self.update_idletasks()
        x = (self.winfo_screenwidth() // 2) - 175
        y = (self.winfo_screenheight() // 2) - 100
        self.geometry(f"350x200+{x}+{y}")
        
        self.create_widgets()
        self.wait_window(self)
    
    def create_widgets(self):
        frame = ttk.Frame(self, padding=20)
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="用户名:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.username = ttk.Entry(frame, width=30)
        self.username.grid(row=0, column=1, pady=5)
        
        ttk.Label(frame, text="密码:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.password = ttk.Entry(frame, width=30, show="•")
        self.password.grid(row=1, column=1, pady=5)
        
        ttk.Label(frame, text="可用账户:", foreground="gray").grid(row=2, column=0, columnspan=2, pady=(10, 0))
        ttk.Label(frame, text="pm/pm123 | dev/dev123 | qa/qa123 | admin/admin123", 
                 foreground="gray", font=("TkDefaultFont", 8)).grid(row=3, column=0, columnspan=2)
        
        btn_frame = ttk.Frame(frame)
        btn_frame.grid(row=4, column=0, columnspan=2, pady=15)
        
        ttk.Button(btn_frame, text="登录", command=self.login).pack(side=tk.LEFT, padx=10)
        ttk.Button(btn_frame, text="取消", command=self.cancel).pack(side=tk.LEFT, padx=10)
    
    def login(self):
        username = self.username.get().strip()
        password = self.password.get().strip()
        
        if not username or not password:
            messagebox.showerror("错误", "请输入用户名和密码")
            return
        
        if username in USER_CREDENTIALS and USER_CREDENTIALS[username] == password:
            self.result = {'username': username, 'role': 'admin'}
            self.destroy()
        else:
            messagebox.showerror("错误", "用户名或密码错误")
    
    def cancel(self):
        self.destroy()


class AssetDialog(tk.Toplevel):
    """资产新增/编辑对话框"""
    
    def __init__(self, parent, asset: FixedAsset = None, is_edit: bool = False):
        super().__init__(parent)
        self.asset = asset
        self.is_edit = is_edit
        self.result = None
        
        self.title("编辑资产" if is_edit else "新增资产")
        self.geometry("600x500")
        self.transient(parent)
        self.grab_set()
        
        self.update_idletasks()
        x = (self.winfo_screenwidth() // 2) - 300
        y = (self.winfo_screenheight() // 2) - 250
        self.geometry(f"600x500+{x}+{y}")
        
        self.create_widgets()
        self.wait_window(self)
    
    def create_widgets(self):
        frame = ttk.Frame(self, padding=20)
        frame.pack(fill=tk.BOTH, expand=True)
        
        # 第一行
        ttk.Label(frame, text="资产名称:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.name = ttk.Entry(frame, width=40)
        self.name.grid(row=0, column=1, pady=5)
        
        # 第二行
        ttk.Label(frame, text="资产类别:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.category = ttk.Combobox(frame, width=37, state="readonly")
        self.category['values'] = list(DEPRECIATION_YEARS_MAP.keys())
        self.category.grid(row=1, column=1, pady=5)
        
        # 第三行
        ttk.Label(frame, text="原值 (元):").grid(row=2, column=0, sticky=tk.W, pady=5)
        self.original_value = ttk.Entry(frame, width=40)
        self.original_value.grid(row=2, column=1, pady=5)
        
        # 第四行
        ttk.Label(frame, text="购置日期:").grid(row=3, column=0, sticky=tk.W, pady=5)
        self.purchase_date = ttk.Entry(frame, width=40)
        self.purchase_date.grid(row=3, column=1, pady=5)
        ttk.Label(frame, text="格式：YYYY-MM-DD", foreground="gray").grid(row=3, column=2, sticky=tk.W, padx=5)
        
        # 第五行
        ttk.Label(frame, text="使用部门:").grid(row=4, column=0, sticky=tk.W, pady=5)
        self.department = ttk.Entry(frame, width=40)
        self.department.grid(row=4, column=1, pady=5)
        
        # 第六行
        ttk.Label(frame, text="使用人:").grid(row=5, column=0, sticky=tk.W, pady=5)
        self.user = ttk.Entry(frame, width=40)
        self.user.grid(row=5, column=1, pady=5)
        
        # 第七行
        ttk.Label(frame, text="存放地点:").grid(row=6, column=0, sticky=tk.W, pady=5)
        self.location = ttk.Entry(frame, width=40)
        self.location.grid(row=6, column=1, pady=5)
        
        # 第八行
        ttk.Label(frame, text="资产状态:").grid(row=7, column=0, sticky=tk.W, pady=5)
        self.status = ttk.Combobox(frame, width=37, state="readonly")
        self.status['values'] = VALID_STATUS
        self.status.grid(row=7, column=1, pady=5)
        
        # 第九行
        ttk.Label(frame, text="备注:").grid(row=8, column=0, sticky=tk.W, pady=5)
        self.remark = ttk.Entry(frame, width=40)
        self.remark.grid(row=8, column=1, pady=5)
        
        # 按钮
        btn_frame = ttk.Frame(frame)
        btn_frame.grid(row=9, column=0, columnspan=2, pady=20)
        
        ttk.Button(btn_frame, text="保存", command=self.save).pack(side=tk.LEFT, padx=10)
        ttk.Button(btn_frame, text="取消", command=self.cancel).pack(side=tk.LEFT, padx=10)
        
        # 填充数据
        if self.asset:
            self.name.insert(0, self.asset.name)
            self.category.set(self.asset.category)
            self.original_value.insert(0, str(self.asset.original_value))
            self.purchase_date.insert(0, self.asset.purchase_date)
            self.department.insert(0, self.asset.department)
            self.user.insert(0, self.asset.user)
            self.location.insert(0, self.asset.location)
            self.status.set(self.asset.status)
            self.remark.insert(0, self.asset.remark)
    
    def save(self):
        try:
            name = self.name.get().strip()
            category = self.category.get().strip()
            original_value = float(self.original_value.get().strip())
            purchase_date = self.purchase_date.get().strip()
            
            if not name or not category or not purchase_date:
                messagebox.showerror("错误", "请填写必填项")
                return
            
            asset = FixedAsset(
                asset_id=self.asset.asset_id if self.asset else "",
                name=name,
                category=category,
                original_value=original_value,
                purchase_date=purchase_date,
                department=self.department.get().strip(),
                user=self.user.get().strip(),
                location=self.location.get().strip(),
                status=self.status.get().strip() or "在用",
                remark=self.remark.get().strip()
            )
            
            self.result = asset
            self.destroy()
        except ValueError:
            messagebox.showerror("错误", "原值必须为数字")
    
    def cancel(self):
        self.destroy()


# ==================== 主界面 ====================

class AssetManagerApp:
    """固定资产管理系统主界面 V11.0 - 符合开发说明书 2.1 节"""
    
    def __init__(self, root):
        self.root = root
        self.root.title(f"工会固定资产管理系统 V{VERSION}")
        self.root.geometry("1200x700")
        
        self.db = DatabaseManager()
        self.current_user = None
        self.login_time = None
        
        self.assets = []
        self.filtered_assets = []
        
        self.create_login()
    
    def create_login(self):
        """创建登录界面"""
        self.login_frame = ttk.Frame(self.root, padding=20)
        self.login_frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(self.login_frame, text="工会固定资产管理系统", font=("Arial", 16, "bold")).pack(pady=20)
        ttk.Label(self.login_frame, text=f"V{VERSION} - 工会会计制度合规版", font=("Arial", 12), foreground="blue").pack(pady=5)
        
        btn_frame = ttk.Frame(self.login_frame)
        btn_frame.pack(pady=20)
        
        ttk.Button(btn_frame, text="用户登录", command=self.login).pack(side=tk.LEFT, padx=10)
        ttk.Button(btn_frame, text="退出", command=self.root.quit).pack(side=tk.LEFT, padx=10)
    
    def login(self):
        dialog = LoginDialog(self.root)
        if dialog.result:
            self.current_user = dialog.result['username']
            self.login_time = time.time()
            self.login_frame.pack_forget()
            self.create_main_interface()
            self.refresh_assets()
    
    def check_session(self):
        """检查会话超时 - 符合开发说明书 5.1 节"""
        if self.login_time and (time.time() - self.login_time) > 3600:
            messagebox.showwarning("会话超时", "登录已超时，请重新登录")
            self.current_user = None
            self.login_time = None
            self.main_frame.pack_forget()
            self.create_login()
            return False
        return True
    
    def create_main_interface(self):
        """创建主界面 - 符合开发说明书 2.1 节"""
        self.main_frame = ttk.Frame(self.root, padding=10)
        self.main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 顶部信息栏
        info_frame = ttk.Frame(self.main_frame)
        info_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(info_frame, text=f"当前用户：{self.current_user}", font=("Arial", 10, "bold")).pack(side=tk.LEFT)
        ttk.Label(info_frame, text=f"  |  V{VERSION} 工会会计制度合规版", foreground="blue").pack(side=tk.LEFT, padx=10)
        
        ttk.Button(info_frame, text="退出登录", command=self.logout).pack(side=tk.RIGHT)
        
        # 工具栏
        toolbar = ttk.Frame(self.main_frame)
        toolbar.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Button(toolbar, text="新增资产", command=self.add_asset).pack(side=tk.LEFT, padx=2)
        ttk.Button(toolbar, text="编辑资产", command=self.edit_asset).pack(side=tk.LEFT, padx=2)
        ttk.Button(toolbar, text="删除资产", command=self.delete_asset).pack(side=tk.LEFT, padx=2)
        ttk.Button(toolbar, text="批量删除", command=self.delete_batch).pack(side=tk.LEFT, padx=2)
        
        ttk.Separator(toolbar, orient=tk.VERTICAL).pack(side=tk.LEFT, padx=10, fill=tk.Y)
        
        ttk.Button(toolbar, text="下载导入模板", command=self.download_template).pack(side=tk.LEFT, padx=2)
        ttk.Button(toolbar, text="导入 Excel", command=self.import_excel).pack(side=tk.LEFT, padx=2)
        ttk.Button(toolbar, text="导出选中", command=self.export_selected).pack(side=tk.LEFT, padx=2)
        ttk.Button(toolbar, text="导出全部", command=self.export_all).pack(side=tk.LEFT, padx=2)
        
        ttk.Separator(toolbar, orient=tk.VERTICAL).pack(side=tk.LEFT, padx=10, fill=tk.Y)
        
        ttk.Button(toolbar, text="折旧查询", command=self.show_depreciation).pack(side=tk.LEFT, padx=2)
        ttk.Button(toolbar, text="月度折旧表", command=self.show_monthly_depreciation).pack(side=tk.LEFT, padx=2)
        ttk.Button(toolbar, text="审计日志", command=self.show_audit_logs).pack(side=tk.LEFT, padx=2)
        
        # 搜索栏
        search_frame = ttk.Frame(self.main_frame)
        search_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(search_frame, text="搜索:").pack(side=tk.LEFT)
        self.search_var = tk.StringVar()
        self.search_var.trace_add('write', lambda *args: self.filter_assets())
        search_entry = ttk.Entry(search_frame, textvariable=self.search_var, width=40)
        search_entry.pack(side=tk.LEFT, padx=5)
        
        ttk.Label(search_frame, text="类别:").pack(side=tk.LEFT, padx=(10, 0))
        self.category_filter = ttk.Combobox(search_frame, width=20, state="readonly")
        self.category_filter['values'] = ['全部'] + list(set(DEPRECIATION_YEARS_MAP.keys()))
        self.category_filter.set('全部')
        self.category_filter.bind('<<ComboboxSelected>>', lambda e: self.filter_assets())
        self.category_filter.pack(side=tk.LEFT, padx=5)
        
        ttk.Label(search_frame, text="状态:").pack(side=tk.LEFT, padx=(10, 0))
        self.status_filter = ttk.Combobox(search_frame, width=10, state="readonly")
        self.status_filter['values'] = ['全部'] + VALID_STATUS
        self.status_filter.set('全部')
        self.status_filter.bind('<<ComboboxSelected>>', lambda e: self.filter_assets())
        self.status_filter.pack(side=tk.LEFT, padx=5)
        
        # 资产列表
        list_frame = ttk.Frame(self.main_frame)
        list_frame.pack(fill=tk.BOTH, expand=True)
        
        columns = ('asset_id', 'name', 'category', 'original_value', 'purchase_date', 
                   'department', 'user', 'location', 'status', 'remark')
        
        self.tree = ttk.Treeview(list_frame, columns=columns, show='headings', selectmode='extended')
        
        self.tree.heading('asset_id', text='资产编号')
        self.tree.heading('name', text='资产名称')
        self.tree.heading('category', text='资产类别')
        self.tree.heading('original_value', text='原值 (元)')
        self.tree.heading('purchase_date', text='购置日期')
        self.tree.heading('department', text='使用部门')
        self.tree.heading('user', text='使用人')
        self.tree.heading('location', text='存放地点')
        self.tree.heading('status', text='状态')
        self.tree.heading('remark', text='备注')
        
        self.tree.column('asset_id', width=160)
        self.tree.column('name', width=120)
        self.tree.column('category', width=150)
        self.tree.column('original_value', width=90, anchor=tk.E)
        self.tree.column('purchase_date', width=90, anchor=tk.CENTER)
        self.tree.column('department', width=100)
        self.tree.column('user', width=80)
        self.tree.column('location', width=100)
        self.tree.column('status', width=70, anchor=tk.CENTER)
        self.tree.column('remark', width=100)
        
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 状态栏
        self.status_var = tk.StringVar()
        self.status_var.set("就绪")
        status_bar = ttk.Label(self.main_frame, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W)
        status_bar.pack(fill=tk.X, pady=(10, 0))
        
        # 首页净值展示
        self.net_value_label = ttk.Label(self.main_frame, text="", font=("Arial", 10, "bold"), foreground="blue")
        self.net_value_label.pack(anchor=tk.E, pady=5)
        
        # 绑定刷新
        self.root.after(1000, self.check_session)
    
    def logout(self):
        """退出登录"""
        self.current_user = None
        self.login_time = None
        self.main_frame.pack_forget()
        self.create_login()
    
    def refresh_assets(self):
        """刷新资产列表"""
        self.assets = self.db.get_all_assets()
        self.filter_assets()
    
    def filter_assets(self):
        """筛选资产"""
        search = self.search_var.get().lower()
        category = self.category_filter.get()
        status = self.status_filter.get()
        
        self.filtered_assets = []
        for asset in self.assets:
            if search:
                if not any(search in str(getattr(asset, f)).lower() 
                          for f in ['asset_id', 'name', 'category', 'department', 'user', 'location']):
                    continue
            
            if category != '全部' and asset.category != category:
                continue
            
            if status != '全部' and asset.status != status:
                continue
            
            self.filtered_assets.append(asset)
        
        self.update_tree()
        self.update_net_value()
    
    def update_tree(self):
        """更新树形列表"""
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        for asset in self.filtered_assets:
            self.tree.insert('', tk.END, values=(
                asset.asset_id, asset.name, asset.category, 
                f"{asset.original_value:,.2f}", asset.purchase_date,
                asset.department, asset.user, asset.location, asset.status, asset.remark
            ), tags=(asset.asset_id,))
        
        self.status_var.set(f"共 {len(self.filtered_assets)} 项资产")
    
    def update_net_value(self):
        """更新净值展示"""
        total_net = sum(
            DepreciationCalculator.calculate(asset).net_value 
            for asset in self.filtered_assets
        )
        self.net_value_label.config(text=f"当前净值合计：¥{total_net:,.2f}")
    
    def add_asset(self):
        """新增资产"""
        if not self.check_session():
            return
        
        dialog = AssetDialog(self.root)
        if dialog.result:
            existing_codes = self.db.get_asset_codes()
            asset_code, warning = generate_asset_code(dialog.result.purchase_date, dialog.result.category, existing_codes)
            
            if warning:
                messagebox.showwarning("警告", warning)
            
            dialog.result.asset_id = asset_code
            
            if self.db.add_asset(dialog.result):
                self.refresh_assets()
                self.status_var.set(f"已添加资产：{asset_code}")
                messagebox.showinfo("成功", f"资产已添加\n编号：{asset_code}")
            else:
                messagebox.showerror("错误", "添加失败，资产编号可能已存在")
    
    def edit_asset(self):
        """编辑资产"""
        if not self.check_session():
            return
        
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("警告", "请选择要编辑的资产")
            return
        
        item = self.tree.item(selection[0])
        asset_id = item['values'][0]
        asset = self.db.get_asset_by_id(asset_id)
        
        if asset:
            dialog = AssetDialog(self.root, asset, is_edit=True)
            if dialog.result:
                dialog.result.asset_id = asset_id
                if self.db.update_asset(dialog.result):
                    self.refresh_assets()
                    self.status_var.set(f"已更新资产：{asset_id}")
                else:
                    messagebox.showerror("错误", "更新失败")
    
    def delete_asset(self):
        """删除资产"""
        if not self.check_session():
            return
        
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("警告", "请选择要删除的资产")
            return
        
        if messagebox.askyesno("确认", "确定要删除选中的资产吗？"):
            item = self.tree.item(selection[0])
            asset_id = item['values'][0]
            
            if self.db.delete_asset(asset_id):
                self.refresh_assets()
                self.status_var.set(f"已删除资产：{asset_id}")
    
    def delete_batch(self):
        """批量删除"""
        if not self.check_session():
            return
        
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("警告", "请选择要删除的资产")
            return
        
        if messagebox.askyesno("确认", f"确定要删除选中的 {len(selection)} 项资产吗？"):
            asset_ids = [self.tree.item(item)['values'][0] for item in selection]
            count = self.db.delete_assets_batch(asset_ids)
            self.refresh_assets()
            self.status_var.set(f"已批量删除 {count} 项资产")
    
    def download_template(self):
        """下载 Excel 导入模板 - V11.0 新增"""
        if not self.check_session():
            return
        
        if not EXCEL_AVAILABLE:
            messagebox.showerror("错误", "未安装 openpyxl 库，无法生成模板")
            return
        
        file_path = filedialog.asksaveasfilename(
            title="保存导入模板",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx")],
            initialfile="固定资产导入模板.xlsx"
        )
        
        if file_path:
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "导入模板"
                
                # 表头 - 符合开发说明书 6.1 节
                headers = ['资产编号', '资产名称', '资产类别', '原值(元)', '购置日期', 
                          '使用部门', '使用人', '存放地点', '状态', '备注']
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                
                # 示例数据
                example_data = [
                    ['', '联想电脑', '通用设备 - 计算机设备', 5000.00, '2026-03-25', '财务部', '张三', '办公室', '在用', '新购置'],
                    ['', '办公桌', '家具、用具及装具 - 家具', 800.00, '2026-03-20', '行政部', '李四', '会议室', '在用', ''],
                ]
                
                for row_idx, row_data in enumerate(example_data, 2):
                    for col_idx, value in enumerate(row_data, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)
                
                # 调整列宽
                column_widths = [15, 20, 30, 12, 12, 15, 10, 15, 10, 20]
                for i, width in enumerate(column_widths, 1):
                    ws.column_dimensions[get_column_letter(i)].width = width
                
                # 添加说明工作表
                ws_help = wb.create_sheet("填写说明")
                help_text = [
                    ["字段", "说明", "示例"],
                    ["资产编号", "留空，系统自动生成", ""],
                    ["资产名称", "必填，资产的中文名称", "联想电脑"],
                    ["资产类别", "必填，从预设类别中选择", "通用设备 - 计算机设备"],
                    ["原值", "必填，购置时的金额（元）", "5000.00"],
                    ["购置日期", "必填，格式：YYYY-MM-DD", "2026-03-25"],
                    ["使用部门", "选填，使用资产的部门", "财务部"],
                    ["使用人", "选填，资产负责人", "张三"],
                    ["存放地点", "选填，资产存放位置", "办公室"],
                    ["状态", "选填，在用/闲置/待维修/待报废", "在用"],
                    ["备注", "选填，其他说明信息", "新购置"],
                    ["", "", ""],
                    ["注意事项：", "", ""],
                    ["1. 资产编号留空，系统会自动生成（格式：国标分类码-YY-MM-DD-XXX，依据：GB/T 14885-2022）", "", ""],
                    ["2. 日期支持多种格式，会自动清洗为标准格式", "", ""],
                    ["3. 资产类别必须是预设类别之一，详见下方【资产类别列表】", "", ""],
                    ["4. 备注字段可以包含任意文本内容", "", ""],
                    ["", "", ""],
                    ["【资产类别列表】", "", ""],
                    ["", "", ""],
                    ["一、房屋及构筑物", "折旧年限", ""],
                    ["房屋及构筑物 - 钢结构", "600个月(50年)", ""],
                    ["房屋及构筑物 - 钢筋混凝土结构", "600个月(50年)", ""],
                    ["房屋及构筑物 - 砖混结构", "360个月(30年)", ""],
                    ["房屋及构筑物 - 砖木结构", "360个月(30年)", ""],
                    ["房屋及构筑物 - 简易房", "96个月(8年)", ""],
                    ["房屋及构筑物 - 房屋附属设施", "96个月(8年)", ""],
                    ["房屋及构筑物 - 构筑物", "96个月(8年)", ""],
                    ["", "", ""],
                    ["二、通用设备", "折旧年限", ""],
                    ["通用设备 - 计算机设备", "72个月(6年)", ""],
                    ["通用设备 - 办公设备", "72个月(6年)", ""],
                    ["通用设备 - 车辆", "96个月(8年)", ""],
                    ["通用设备 - 图书档案设备", "60个月(5年)", ""],
                    ["通用设备 - 机械设备", "120个月(10年)", ""],
                    ["通用设备 - 电气设备", "60个月(5年)", ""],
                    ["通用设备 - 雷达、无线电和卫星导航设备", "120个月(10年)", ""],
                    ["通用设备 - 通信设备", "60个月(5年)", ""],
                    ["通用设备 - 广播、电视、电影设备", "60个月(5年)", ""],
                    ["通用设备 - 仪器仪表", "60个月(5年)", ""],
                    ["通用设备 - 电子和通信测量设备", "60个月(5年)", ""],
                    ["通用设备 - 计量标准器具及量具、衡器", "60个月(5年)", ""],
                    ["", "", ""],
                    ["三、专用设备", "折旧年限", ""],
                    ["专用设备 - 食品加工专用设备", "120个月(10年)", ""],
                    ["专用设备 - 纺织设备", "120个月(10年)", ""],
                    ["专用设备 - 缝纫、服饰、制革和毛皮加工设备", "120个月(10年)", ""],
                    ["专用设备 - 医疗设备", "60个月(5年)", ""],
                    ["专用设备 - 安全生产设备", "120个月(10年)", ""],
                    ["专用设备 - 环境污染防治设备", "120个月(10年)", ""],
                    ["专用设备 - 文艺设备", "60个月(5年)", ""],
                    ["专用设备 - 体育设备", "60个月(5年)", ""],
                    ["专用设备 - 娱乐设备", "60个月(5年)", ""],
                    ["", "", ""],
                    ["四、家具、用具及装具", "折旧年限", ""],
                    ["家具、用具及装具 - 家具", "180个月(15年)", ""],
                    ["家具、用具及装具 - 用具、装具", "60个月(5年)", ""],
                ]
                
                for row_idx, row_data in enumerate(help_text, 1):
                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws_help.cell(row=row_idx, column=col_idx, value=value)
                        if row_idx == 1:
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                            cell.font = Font(bold=True, color="FFFFFF")
                
                ws_help.column_dimensions['A'].width = 15
                ws_help.column_dimensions['B'].width = 50
                ws_help.column_dimensions['C'].width = 20
                
                wb.save(file_path)
                messagebox.showinfo("成功", f"模板已保存到：{file_path}\n\n请按照'填写说明'工作表的指引填写数据后导入。")
                self.status_var.set("导入模板已下载")
            except Exception as e:
                messagebox.showerror("错误", f"生成模板失败：{e}")
    
    def import_excel(self):
        """导入 Excel - 修复数据库锁定问题，使用批量事务"""
        if not self.check_session():
            return
        
        if not EXCEL_AVAILABLE:
            messagebox.showerror("错误", "未安装 openpyxl 库，无法导入 Excel")
            return
        
        file_path = filedialog.askopenfilename(
            title="选择 Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx *.xls")]
        )
        
        if not file_path:
            return
        
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path)
            ws = wb.active
            
            # 预读取所有行数据
            rows_data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                # 检查行是否有任何有效数据（资产名称或原值不为空）
                if row and len(row) >= 3 and (row[0] or row[1] or row[2]):
                    rows_data.append(row)
            
            if not rows_data:
                messagebox.showwarning("警告", "未找到有效数据：Excel 文件中数据行均为空或格式不正确，请检查数据是否从第2行开始填写，并确保资产编号、资产名称或原值列有数据")
                return
            
            # 批量处理：使用单一事务
            conn = self.db.get_connection()
            cursor = conn.cursor()
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            existing_codes = self.db.get_asset_codes()
            
            count = 0
            failed_rows = []
            
            try:
                for row_idx, row in enumerate(rows_data, start=2):
                    try:
                        # 日期清洗
                        purchase_date = str(row[4]) if row[4] else datetime.now().strftime("%Y-%m-%d")
                        if ' ' in purchase_date:
                            purchase_date = purchase_date.split()[0]
                        
                        # 生成资产编码（根据类别自动生成国标编码）
                        category = str(row[2]) if row[2] else "通用设备 - 办公设备"
                        asset_code, warning = generate_asset_code(purchase_date, category, existing_codes)
                        existing_codes.append(asset_code)
                        
                        # 插入数据（参数化查询）
                        cursor.execute('''
                            INSERT INTO assets (asset_id, name, category, original_value, purchase_date,
                                               department, user, location, status, remark, created_time, updated_time)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        ''', (
                            asset_code,
                            str(row[1]) if row[1] else "未命名",
                            str(row[2]) if row[2] else "通用设备 - 办公设备",
                            float(row[3]) if row[3] else 0.0,
                            purchase_date,
                            str(row[5]) if row[5] else "",
                            str(row[6]) if row[6] else "",
                            str(row[7]) if row[7] else "",
                            str(row[8]) if row[8] else "在用",
                            str(row[9]) if row[9] else "",
                            now, now
                        ))
                        
                        # 记录审计日志
                        cursor.execute('''
                            INSERT INTO audit_logs (operation_type, asset_id, user_id, old_values, new_values, operation_time)
                            VALUES (?, ?, ?, ?, ?, ?)
                        ''', ('CREATE', asset_code, 'admin', None, None, now))
                        
                        count += 1
                        
                        # 每100条提交一次，避免事务过大
                        if count % 100 == 0:
                            conn.commit()
                            
                    except Exception as e:
                        failed_rows.append(f"第{row_idx}行: {str(e)}")
                        continue
                
                # 最终提交
                conn.commit()
                
            except Exception as e:
                conn.rollback()
                raise e
            finally:
                conn.close()
            
            # 刷新界面
            self.refresh_assets()
            
            # 显示结果
            if failed_rows:
                messagebox.showwarning("部分导入成功", 
                    f"成功导入 {count} 项资产\n\n以下行导入失败:\n" + "\n".join(failed_rows[:10]))
            else:
                messagebox.showinfo("成功", f"成功导入 {count} 项资产")
            
            self.status_var.set(f"已导入 {count} 项资产")
            
        except Exception as e:
            messagebox.showerror("错误", f"导入失败：{e}")
    
    def export_selected(self):
        """导出选中资产"""
        if not self.check_session():
            return
        
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("警告", "请选择要导出的资产")
            return
        
        self._export_to_excel([self.tree.item(item)['values'] for item in selection], "选中资产")
    
    def export_all(self):
        """导出全部资产"""
        if not self.check_session():
            return
        
        self._export_to_excel([
            (asset.asset_id, asset.name, asset.category, f"{asset.original_value:,.2f}",
             asset.purchase_date, asset.department, asset.user, asset.location, asset.status, asset.remark)
            for asset in self.filtered_assets
        ], "全部资产")
    
    def _export_to_excel(self, data, title):
        """导出到 Excel - 符合开发说明书 6.2 节"""
        if not EXCEL_AVAILABLE:
            messagebox.showerror("错误", "未安装 openpyxl 库，无法导出 Excel")
            return
        
        file_path = filedialog.asksaveasfilename(
            title="保存 Excel 文件",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx")]
        )
        
        if file_path:
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = title
                
                headers = ['资产编号', '资产名称', '资产类别', '原值 (元)', '购置日期',
                          '使用部门', '使用人', '存放地点', '状态', '备注']
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                
                for row_idx, row_data in enumerate(data, 2):
                    for col_idx, value in enumerate(row_data, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)
                
                wb.save(file_path)
                messagebox.showinfo("成功", f"已导出到：{file_path}")
                self.status_var.set(f"已导出 {len(data)} 项资产")
            except Exception as e:
                messagebox.showerror("错误", f"导出失败：{e}")
    
    def show_depreciation(self):
        """显示折旧查询"""
        if not self.check_session():
            return
        
        dialog = DepreciationDialog(self.root, self.filtered_assets, self.db)
    
    def show_monthly_depreciation(self):
        """显示月度折旧表"""
        if not self.check_session():
            return
        
        dialog = MonthlyDepreciationDialog(self.root, self.filtered_assets, self.db)
    
    def show_audit_logs(self):
        """显示审计日志 - V11.0 新增，符合开发说明书 2.1 节"""
        if not self.check_session():
            return
        
        dialog = AuditLogDialog(self.root, self.db)


class DepreciationDialog(tk.Toplevel):
    """折旧查询对话框 V11.0"""
    
    def __init__(self, parent, assets, db: DatabaseManager):
        super().__init__(parent)
        self.assets = assets
        self.db = db
        self.title("折旧查询")
        self.geometry("1200x650")
        self.transient(parent)
        self.grab_set()
        
        self.update_idletasks()
        x = (self.winfo_screenwidth() // 2) - 600
        y = (self.winfo_screenheight() // 2) - 325
        self.geometry(f"1200x650+{x}+{y}")
        
        self.create_widgets()
    
    def create_widgets(self):
        frame = ttk.Frame(self, padding=10)
        frame.pack(fill=tk.BOTH, expand=True)
        
        # 日期选择
        date_frame = ttk.Frame(frame)
        date_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(date_frame, text="查询日期:").pack(side=tk.LEFT)
        self.date_var = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        date_entry = ttk.Entry(date_frame, textvariable=self.date_var, width=15)
        date_entry.pack(side=tk.LEFT, padx=5)
        
        ttk.Button(date_frame, text="查询", command=self.refresh).pack(side=tk.LEFT, padx=5)
        ttk.Button(date_frame, text="导出 Excel", command=self.export).pack(side=tk.LEFT, padx=5)
        ttk.Button(date_frame, text="生成凭证", command=self.generate_entry).pack(side=tk.LEFT, padx=5)
        ttk.Button(date_frame, text="查看审计日志", command=self.show_audit_log).pack(side=tk.LEFT, padx=5)
        ttk.Button(date_frame, text="关闭", command=self.destroy).pack(side=tk.RIGHT)
        
        # 汇总信息
        self.summary_var = tk.StringVar()
        ttk.Label(frame, textvariable=self.summary_var, font=("Arial", 10, "bold"), 
                 foreground="blue").pack(anchor=tk.W, pady=5)
        
        # 会计分录
        self.entry_var = tk.StringVar()
        entry_label = ttk.Label(frame, textvariable=self.entry_var, font=("Courier", 10), 
                               foreground="green", justify=tk.LEFT)
        entry_label.pack(anchor=tk.W, pady=5)
        
        # 列表 V11.0：新增折旧差额列
        columns = ('asset_id', 'asset_name', 'category', 'original_value', 
                   'monthly_depreciation', 'accumulated_depreciation', 'net_value',
                   'depreciable_months', 'depreciation_status', 'current_month_depreciation',
                   'depreciation_diff')
        
        self.tree = ttk.Treeview(frame, columns=columns, show='headings', selectmode='extended')
        
        self.tree.heading('asset_id', text='资产编号')
        self.tree.heading('asset_name', text='资产名称')
        self.tree.heading('category', text='资产类别')
        self.tree.heading('original_value', text='原值')
        self.tree.heading('monthly_depreciation', text='月折旧额')
        self.tree.heading('accumulated_depreciation', text='累计折旧')
        self.tree.heading('net_value', text='净值')
        self.tree.heading('depreciable_months', text='已计提月数')
        self.tree.heading('depreciation_status', text='状态')
        self.tree.heading('current_month_depreciation', text='本月计提')
        self.tree.heading('depreciation_diff', text='折旧差额')
        
        self.tree.column('asset_id', width=160)
        self.tree.column('asset_name', width=120)
        self.tree.column('category', width=150)
        self.tree.column('original_value', width=90, anchor=tk.E)
        self.tree.column('monthly_depreciation', width=90, anchor=tk.E)
        self.tree.column('accumulated_depreciation', width=90, anchor=tk.E)
        self.tree.column('net_value', width=90, anchor=tk.E)
        self.tree.column('depreciable_months', width=80, anchor=tk.CENTER)
        self.tree.column('depreciation_status', width=80, anchor=tk.CENTER)
        self.tree.column('current_month_depreciation', width=90, anchor=tk.E)
        self.tree.column('depreciation_diff', width=90, anchor=tk.E)
        
        scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.results = []
        self.refresh()
    
    def refresh(self):
        """刷新数据"""
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        query_date = self.date_var.get()
        self.results = []
        
        total_original = 0
        total_monthly = 0
        total_accumulated = 0
        total_net = 0
        total_current_month = 0
        
        for asset in self.assets:
            result = DepreciationCalculator.calculate(asset, query_date)
            self.results.append(result)
            
            # V11.0：记录审计日志 - 符合开发说明书 4.1 节
            self.db.log_depreciation(result, query_date)
            
            total_original += asset.original_value
            total_monthly += result.monthly_depreciation
            total_accumulated += result.accumulated_depreciation
            total_net += result.net_value
            total_current_month += result.current_month_depreciation
            
            self.tree.insert('', tk.END, values=(
                result.asset_id, result.asset_name, result.category,
                f"{result.original_value:,.2f}", f"{result.monthly_depreciation:,.2f}",
                f"{result.accumulated_depreciation:,.2f}", f"{result.net_value:,.2f}",
                result.depreciable_months, result.depreciation_status,
                f"{result.current_month_depreciation:,.2f}",
                f"{result.depreciation_diff:,.2f}"
            ))
        
        self.summary_var.set(
            f"原值合计：¥{total_original:,.2f} | "
            f"月折旧额：¥{total_monthly:,.2f} | "
            f"累计折旧：¥{total_accumulated:,.2f} | "
            f"净值合计：¥{total_net:,.2f} | "
            f"本月计提：¥{total_current_month:,.2f}"
        )
        
        # 生成会计分录
        self.entry_var.set(DepreciationCalculator.generate_accounting_entry(self.results))
    
    def generate_entry(self):
        """生成并显示会计分录 - 符合开发说明书 4.1 节"""
        entry = DepreciationCalculator.generate_accounting_entry(self.results)
        messagebox.showinfo("会计分录", entry)
    
    def show_audit_log(self):
        """显示审计日志"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("警告", "请选择要查看的资产")
            return
        
        item = self.tree.item(selection[0])
        asset_id = item['values'][0]
        
        # 查找对应结果
        for result in self.results:
            if result.asset_id == asset_id:
                messagebox.showinfo("审计日志", result.audit_log)
                return
    
    def export(self):
        """导出 Excel"""
        if not EXCEL_AVAILABLE:
            messagebox.showerror("错误", "未安装 openpyxl 库")
            return
        
        file_path = filedialog.asksaveasfilename(
            title="保存 Excel 文件",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx")]
        )
        
        if file_path:
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "折旧查询"
                
                headers = ['资产编号', '资产名称', '资产类别', '原值', '月折旧额',
                          '累计折旧', '净值', '已计提月数', '状态', '本月计提', '折旧差额']
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                
                for row_idx, item in enumerate(self.tree.get_children(), 2):
                    values = self.tree.item(item)['values']
                    for col_idx, value in enumerate(values, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)
                
                wb.save(file_path)
                messagebox.showinfo("成功", f"已导出到：{file_path}")
            except Exception as e:
                messagebox.showerror("错误", f"导出失败：{e}")


class MonthlyDepreciationDialog(tk.Toplevel):
    """月度折旧表对话框 V11.0"""
    
    def __init__(self, parent, assets, db: DatabaseManager):
        super().__init__(parent)
        self.assets = assets
        self.db = db
        self.title("月度折旧表")
        self.geometry("1200x650")
        self.transient(parent)
        self.grab_set()
        
        self.update_idletasks()
        x = (self.winfo_screenwidth() // 2) - 600
        y = (self.winfo_screenheight() // 2) - 325
        self.geometry(f"1200x650+{x}+{y}")
        
        self.create_widgets()
    
    def create_widgets(self):
        frame = ttk.Frame(self, padding=10)
        frame.pack(fill=tk.BOTH, expand=True)
        
        # 日期选择
        date_frame = ttk.Frame(frame)
        date_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(date_frame, text="查询月份:").pack(side=tk.LEFT)
        self.month_var = tk.StringVar(value=datetime.now().strftime("%Y-%m"))
        month_entry = ttk.Entry(date_frame, textvariable=self.month_var, width=15)
        month_entry.pack(side=tk.LEFT, padx=5)
        ttk.Label(date_frame, text="格式：YYYY-MM", foreground="gray").pack(side=tk.LEFT)
        
        ttk.Button(date_frame, text="查询", command=self.refresh).pack(side=tk.LEFT, padx=5)
        ttk.Button(date_frame, text="导出 Excel", command=self.export).pack(side=tk.LEFT, padx=5)
        ttk.Button(date_frame, text="关闭", command=self.destroy).pack(side=tk.RIGHT)
        
        # 汇总信息
        self.summary_var = tk.StringVar()
        ttk.Label(frame, textvariable=self.summary_var, font=("Arial", 10, "bold"), 
                 foreground="blue").pack(anchor=tk.W, pady=5)
        
        # 列表 V11.0
        columns = ('asset_id', 'asset_name', 'category', 'department', 'user', 'location',
                   'status', 'original_value', 'monthly_depreciation', 'depreciation_diff',
                   'current_month_total', 'purchase_date', 'depreciable_months')
        
        self.tree = ttk.Treeview(frame, columns=columns, show='headings', selectmode='extended')
        
        self.tree.heading('asset_id', text='资产编号')
        self.tree.heading('asset_name', text='资产名称')
        self.tree.heading('category', text='资产类别')
        self.tree.heading('department', text='使用部门')
        self.tree.heading('user', text='使用人')
        self.tree.heading('location', text='存放地点')
        self.tree.heading('status', text='状态')
        self.tree.heading('original_value', text='原值')
        self.tree.heading('monthly_depreciation', text='正常折旧额')
        self.tree.heading('depreciation_diff', text='折旧差额')
        self.tree.heading('current_month_total', text='当月总折旧额')
        self.tree.heading('purchase_date', text='购置日期')
        self.tree.heading('depreciable_months', text='已计提折旧月数')
        
        self.tree.column('asset_id', width=160)
        self.tree.column('asset_name', width=120)
        self.tree.column('category', width=150)
        self.tree.column('department', width=100)
        self.tree.column('user', width=80)
        self.tree.column('location', width=100)
        self.tree.column('status', width=70, anchor=tk.CENTER)
        self.tree.column('original_value', width=90, anchor=tk.E)
        self.tree.column('monthly_depreciation', width=90, anchor=tk.E)
        self.tree.column('depreciation_diff', width=90, anchor=tk.E)
        self.tree.column('current_month_total', width=90, anchor=tk.E)
        self.tree.column('purchase_date', width=90, anchor=tk.CENTER)
        self.tree.column('depreciable_months', width=100, anchor=tk.CENTER)
        
        scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.refresh()
    
    def refresh(self):
        """刷新数据"""
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        month_str = self.month_var.get()
        try:
            query_date = f"{month_str}-01"
            query_dt = datetime.strptime(query_date, "%Y-%m-%d")
        except ValueError:
            messagebox.showerror("错误", "月份格式错误，请使用 YYYY-MM 格式")
            return
        
        results = []
        total_original = 0
        total_monthly = 0
        total_diff = 0
        total_current = 0
        
        for asset in self.assets:
            result = DepreciationCalculator.calculate(asset, query_date, for_monthly_report=True)
            results.append(result)
            
            # V11.0：剔除已提足、待报废、待维修资产
            should_show = True
            if result.depreciable_months >= result.depreciation_months:
                should_show = False
            if asset.status in ['待报废', '待维修']:
                should_show = False
            
            if should_show:
                diff = result.depreciation_diff
                current_total = result.current_month_depreciation
                
                total_original += asset.original_value
                total_monthly += result.monthly_depreciation
                total_diff += diff
                total_current += current_total
                
                self.tree.insert('', tk.END, values=(
                    result.asset_id, result.asset_name, result.category,
                    result.department, result.user, result.location,
                    result.status, f"{result.original_value:,.2f}",
                    f"{result.monthly_depreciation:,.2f}", f"{diff:,.2f}",
                    f"{current_total:,.2f}", result.purchase_date,
                    result.depreciable_months
                ))
        
        self.summary_var.set(
            f"原值合计：¥{total_original:,.2f} | "
            f"正常折旧额：¥{total_monthly:,.2f} | "
            f"折旧差额：¥{total_diff:,.2f} | "
            f"当月总折旧额：¥{total_current:,.2f}"
        )
    
    def export(self):
        """导出 Excel V11.0"""
        if not EXCEL_AVAILABLE:
            messagebox.showerror("错误", "未安装 openpyxl 库")
            return
        
        file_path = filedialog.asksaveasfilename(
            title="保存 Excel 文件",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx")]
        )
        
        if file_path:
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "月度折旧表"
                
                headers = ['资产编号', '资产名称', '资产类别', '使用部门', '使用人',
                          '存放地点', '状态', '原值', '正常折旧额', '折旧差额',
                          '当月总折旧额', '购置日期', '已计提折旧月数']
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                
                for row_idx, item in enumerate(self.tree.get_children(), 2):
                    values = self.tree.item(item)['values']
                    for col_idx, value in enumerate(values, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)
                
                wb.save(file_path)
                messagebox.showinfo("成功", f"已导出到：{file_path}")
            except Exception as e:
                messagebox.showerror("错误", f"导出失败：{e}")


class AuditLogDialog(tk.Toplevel):
    """审计日志查看对话框 - V11.0 新增，符合开发说明书 2.1 节"""
    
    def __init__(self, parent, db: DatabaseManager):
        super().__init__(parent)
        self.db = db
        self.title("审计日志")
        self.geometry("1000x600")
        self.transient(parent)
        self.grab_set()
        
        self.update_idletasks()
        x = (self.winfo_screenwidth() // 2) - 500
        y = (self.winfo_screenheight() // 2) - 300
        self.geometry(f"1000x600+{x}+{y}")
        
        self.create_widgets()
    
    def create_widgets(self):
        frame = ttk.Frame(self, padding=10)
        frame.pack(fill=tk.BOTH, expand=True)
        
        # 工具栏
        toolbar = ttk.Frame(frame)
        toolbar.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(toolbar, text="资产编号:").pack(side=tk.LEFT)
        self.asset_id_filter = ttk.Entry(toolbar, width=20)
        self.asset_id_filter.pack(side=tk.LEFT, padx=5)
        
        ttk.Button(toolbar, text="查询", command=self.refresh).pack(side=tk.LEFT, padx=5)
        ttk.Button(toolbar, text="导出", command=self.export).pack(side=tk.LEFT, padx=5)
        ttk.Button(toolbar, text="关闭", command=self.destroy).pack(side=tk.RIGHT)
        
        # 列表
        columns = ('operation_time', 'operation_type', 'asset_id', 'user_id', 'details')
        
        self.tree = ttk.Treeview(frame, columns=columns, show='headings', selectmode='extended')
        
        self.tree.heading('operation_time', text='操作时间')
        self.tree.heading('operation_type', text='操作类型')
        self.tree.heading('asset_id', text='资产编号')
        self.tree.heading('user_id', text='操作人')
        self.tree.heading('details', text='详情')
        
        self.tree.column('operation_time', width=160)
        self.tree.column('operation_type', width=100, anchor=tk.CENTER)
        self.tree.column('asset_id', width=180)
        self.tree.column('user_id', width=100)
        self.tree.column('details', width=400)
        
        scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.refresh()
    
    def refresh(self):
        """刷新审计日志"""
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        asset_id = self.asset_id_filter.get().strip() or None
        logs = self.db.get_audit_logs(asset_id)
        
        for log in logs:
            details = ""
            if log['operation_type'] == 'CREATE':
                details = f"新增资产：{log['new_values'].get('name', '') if log['new_values'] else ''}"
            elif log['operation_type'] == 'UPDATE':
                details = "更新资产信息"
            elif log['operation_type'] == 'DELETE':
                details = "删除资产"
            elif log['operation_type'] == 'DEPRECIATE':
                if log['new_values']:
                    details = f"折旧计算：本月计提={log['new_values'].get('current_month_depreciation', 0):.2f}元"
            
            self.tree.insert('', tk.END, values=(
                log['operation_time'],
                log['operation_type'],
                log['asset_id'],
                log['user_id'],
                details
            ))
    
    def export(self):
        """导出审计日志"""
        if not EXCEL_AVAILABLE:
            messagebox.showerror("错误", "未安装 openpyxl 库")
            return
        
        file_path = filedialog.asksaveasfilename(
            title="保存 Excel 文件",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx")]
        )
        
        if file_path:
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "审计日志"
                
                headers = ['操作时间', '操作类型', '资产编号', '操作人', '详情']
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                
                for row_idx, item in enumerate(self.tree.get_children(), 2):
                    values = self.tree.item(item)['values']
                    for col_idx, value in enumerate(values, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)
                
                wb.save(file_path)
                messagebox.showinfo("成功", f"已导出到：{file_path}")
            except Exception as e:
                messagebox.showerror("错误", f"导出失败：{e}")


# ==================== 主程序入口 ====================

def main():
    root = tk.Tk()
    app = AssetManagerApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
